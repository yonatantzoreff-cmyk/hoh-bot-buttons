"""Tests for Excel parser."""

import tempfile
from datetime import date, datetime, time
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.utils.excel_parser import parse_excel_file, _parse_time_string


def test_parse_time_string():
    """Test time parsing from various formats."""
    assert _parse_time_string("09:30") == time(9, 30)
    assert _parse_time_string("9:30") == time(9, 30)
    assert _parse_time_string("21:00") == time(21, 0)
    assert _parse_time_string("9.30") == time(9, 30)
    assert _parse_time_string("0930") == time(9, 30)
    assert _parse_time_string("930") == time(9, 30)
    assert _parse_time_string("20:30:00") == time(20, 30)
    assert _parse_time_string("31.12.1899 08:05:00") == time(8, 5)
    assert _parse_time_string("") is None
    assert _parse_time_string("invalid") is None
    assert _parse_time_string("25:00") is None  # Invalid hour


def test_parse_excel_file_basic():
    """Test parsing a basic Excel file with Hebrew headers."""
    # Create a temporary Excel file
    wb = Workbook()
    ws = wb.active
    
    # Add Hebrew headers
    ws.append(["תאריך", "שעה", "שם המופע", "שעה טכני", "סדרה", "גוף מבצע / איש קשר", "טלפון", "הערות", "יום"])
    
    # Add a data row
    ws.append([
        date(2025, 12, 25),
        time(20, 0),
        "Test Event",
        time(18, 0),
        "Series A",
        "John Doe",
        "0501234567",
        "Test notes",
        "יום שלישי"
    ])
    
    # Save to temp file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    try:
        # Parse the file
        events = parse_excel_file(tmp_path)
        
        assert len(events) == 1
        event = events[0]
        
        assert event["date"] == date(2025, 12, 25)
        assert event["show_time"] == time(20, 0)
        assert event["name"] == "Test Event"
        assert event["load_in"] == time(18, 0)
        assert event["event_series"] == "Series A"
        assert event["producer_name"] == "John Doe"
        assert event["producer_phone"] == "0501234567"
        assert event["notes"] == "Test notes"
        assert event["row_index"] == 2  # Header is row 1, data starts at row 2
        
    finally:
        # Clean up
        Path(tmp_path).unlink(missing_ok=True)


def test_parse_excel_file_multiple_rows():
    """Test parsing multiple event rows."""
    wb = Workbook()
    ws = wb.active
    
    ws.append(["תאריך", "שעה", "שם המופע", "שעה טכני", "סדרה", "גוף מבצע / איש קשר", "טלפון", "הערות"])
    ws.append([date(2025, 12, 25), time(20, 0), "Event 1", None, None, "Producer 1", "050-1234567", ""])
    ws.append([date(2025, 12, 26), time(19, 30), "Event 2", time(17, 0), "Series B", "Producer 2", "", "Notes 2"])
    ws.append([date(2025, 12, 27), time(21, 0), "Event 3", None, None, None, None, None])
    
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    try:
        events = parse_excel_file(tmp_path)
        
        assert len(events) == 3
        assert events[0]["name"] == "Event 1"
        assert events[1]["name"] == "Event 2"
        assert events[2]["name"] == "Event 3"
        
        # Check optional fields
        assert events[0]["load_in"] is None
        assert events[1]["load_in"] == time(17, 0)
        
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def test_parse_excel_file_empty_rows():
    """Test that empty rows are skipped."""
    wb = Workbook()
    ws = wb.active
    
    ws.append(["תאריך", "שעה", "שם המופע", "שעה טכני", "סדרה", "גוף מבצע / איש קשר", "טלפון", "הערות"])
    ws.append([date(2025, 12, 25), time(20, 0), "Event 1", None, None, "Producer 1", "", ""])
    ws.append([None, None, None, None, None, None, None, None])  # Empty row
    ws.append([date(2025, 12, 26), time(19, 30), "Event 2", None, None, "Producer 2", "", ""])
    
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    try:
        events = parse_excel_file(tmp_path)
        
        # Empty row should be skipped
        assert len(events) == 2
        assert events[0]["name"] == "Event 1"
        assert events[1]["name"] == "Event 2"
        
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def test_parse_excel_file_missing_headers():
    """Test that parser raises error if headers are not found."""
    wb = Workbook()
    ws = wb.active
    
    # Add invalid headers
    ws.append(["Invalid", "Headers", "Here"])
    ws.append([date(2025, 12, 25), time(20, 0), "Event 1"])
    
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    try:
        with pytest.raises(ValueError, match="Could not find valid header row"):
            parse_excel_file(tmp_path)
        
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def test_parse_excel_file_time_strings():
    """Test parsing time values as strings."""
    wb = Workbook()
    ws = wb.active
    
    ws.append(["תאריך", "שעה", "שם המופע", "שעה טכני", "סדרה", "גוף מבצע / איש קשר", "טלפון", "הערות"])
    ws.append([
        "2025-12-25",  # Date as string
        "20:00",        # Time as string
        "Event 1",
        "18:30",        # Time as string
        None,
        "Producer",
        "050-1234567",
        ""
    ])
    
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    try:
        events = parse_excel_file(tmp_path)
        
        assert len(events) == 1
        assert events[0]["date"] == date(2025, 12, 25)
        assert events[0]["show_time"] == time(20, 0)
        assert events[0]["load_in"] == time(18, 30)

    finally:
        Path(tmp_path).unlink(missing_ok=True)


def test_parse_excel_file_with_date_and_time_variants():
    """Ensure parser handles alternate date formats and datetime time values."""
    wb = Workbook()
    ws = wb.active

    ws.append(["תאריך", "שעה", "שם המופע", "שעה טכני"])
    ws.append([
        "01.06.25",  # Short year with dots
        "20:30:00",  # Time string with seconds
        "Event with short date",
        datetime(1899, 12, 31, 18, 45),  # Excel-style datetime for time
    ])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    try:
        events = parse_excel_file(tmp_path)

        assert len(events) == 1
        assert events[0]["date"] == date(2025, 6, 1)
        assert events[0]["show_time"] == time(20, 30)
        assert events[0]["load_in"] == time(18, 45)
    finally:
        Path(tmp_path).unlink(missing_ok=True)
