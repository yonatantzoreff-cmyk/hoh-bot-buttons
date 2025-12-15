"""Excel parser for calendar imports."""

import logging
import re
from datetime import date, datetime, time
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Hebrew column names to internal field mapping
COLUMN_MAPPING = {
    "תאריך": "date",
    "שעה": "show_time",
    "שם המופע": "name",
    "שעה טכני": "load_in",
    "סדרה": "event_series",
    "גוף מבצע / איש קשר": "producer_name",
    "טלפון": "producer_phone",
    "הערות": "notes",
    "יום": "_day",  # Informational only, not stored
}


def parse_excel_file(file_path: str) -> List[Dict[str, Any]]:
    """
    Parse an Excel file and extract event data.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        List of dictionaries with parsed event data
        
    Raises:
        ValueError: If file format is invalid
    """
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        
        if sheet is None:
            raise ValueError("Excel file has no active sheet")
            
        return _parse_sheet(sheet)
        
    except Exception as e:
        logger.error(f"Failed to parse Excel file: {e}")
        raise ValueError(f"Invalid Excel file: {str(e)}")


def _parse_sheet(sheet: Worksheet) -> List[Dict[str, Any]]:
    """Parse a worksheet and extract events."""
    # Find header row and build column mapping
    header_row = None
    col_map = {}
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        for col_idx, cell_value in enumerate(row):
            if cell_value and str(cell_value).strip() in COLUMN_MAPPING:
                if header_row is None:
                    header_row = row_idx
                if header_row == row_idx:
                    field_name = COLUMN_MAPPING[str(cell_value).strip()]
                    col_map[col_idx] = field_name
        
        if header_row is not None:
            break
    
    if not col_map:
        raise ValueError("Could not find valid header row with expected Hebrew column names")
    
    # Parse data rows
    events = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        # Skip empty rows
        if not any(cell for cell in row):
            continue
            
        event_data = {"row_index": row_idx}
        
        for col_idx, cell_value in enumerate(row):
            if col_idx in col_map:
                field_name = col_map[col_idx]
                # Skip informational columns
                if field_name.startswith("_"):
                    continue
                    
                event_data[field_name] = _parse_cell_value(field_name, cell_value)
        
        # Only include rows that have at least a date or name
        if event_data.get("date") or event_data.get("name"):
            events.append(event_data)
    
    return events


def _parse_cell_value(field_name: str, cell_value: Any) -> Optional[Any]:
    """Parse and normalize a cell value based on field type."""
    if cell_value is None or (isinstance(cell_value, str) and not cell_value.strip()):
        return None
    
    try:
        # Date field
        if field_name == "date":
            if isinstance(cell_value, date):
                # Handle both datetime and date objects
                if hasattr(cell_value, 'date'):
                    return cell_value.date()
                return cell_value
            elif isinstance(cell_value, str):
                # Try common date formats
                for fmt in [
                    "%Y-%m-%d",
                    "%d/%m/%Y",
                    "%d/%m/%y",
                    "%d-%m-%Y",
                    "%d-%m-%y",
                    "%d.%m.%Y",
                    "%d.%m.%y",
                ]:
                    try:
                        return datetime.strptime(cell_value.strip(), fmt).date()
                    except ValueError:
                        continue
            return None

        # Time fields
        elif field_name in ("show_time", "load_in"):
            if isinstance(cell_value, datetime):
                return cell_value.time()
            if isinstance(cell_value, time):
                return cell_value
            elif isinstance(cell_value, str):
                return _parse_time_string(cell_value.strip())
            return None
        
        # Text fields
        else:
            return str(cell_value).strip() if cell_value else None
            
    except Exception as e:
        logger.warning(f"Failed to parse {field_name} value '{cell_value}': {e}")
        return None


def _parse_time_string(time_str: str) -> Optional[time]:
    """
    Parse time string in various formats (24h).

    Supports: HH:MM, H:MM, HH:MM:SS, embedded dates with time (e.g. "31.12.1899 20:30:00"),
    HHMM, H.MM, HH.MM
    """
    if not time_str:
        return None

    # Remove whitespace
    time_str = time_str.strip()

    # Extract time component even when embedded in a date string
    colon_match = re.search(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", time_str)
    if colon_match:
        try:
            hours = int(colon_match.group(1))
            minutes = int(colon_match.group(2))
            seconds = int(colon_match.group(3)) if colon_match.group(3) else 0
            if 0 <= hours < 24 and 0 <= minutes < 60 and 0 <= seconds < 60:
                return time(hours, minutes, seconds)
        except ValueError:
            pass

    # Try HH:MM or H:MM format
    if ":" in time_str:
        parts = time_str.split(":")
        if len(parts) in {2, 3}:
            try:
                hours = int(parts[0])
                minutes = int(parts[1])
                seconds = int(parts[2]) if len(parts) == 3 else 0
                if 0 <= hours < 24 and 0 <= minutes < 60 and 0 <= seconds < 60:
                    return time(hours, minutes, seconds)
            except ValueError:
                pass

    # Try H.MM or HH.MM format
    elif "." in time_str:
        parts = time_str.split(".")
        if len(parts) == 2:
            try:
                hours = int(parts[0])
                minutes = int(parts[1])
                if 0 <= hours < 24 and 0 <= minutes < 60:
                    return time(hours, minutes)
            except ValueError:
                pass
    
    # Try HHMM format (4 digits)
    elif len(time_str) == 4 and time_str.isdigit():
        try:
            hours = int(time_str[:2])
            minutes = int(time_str[2:])
            if 0 <= hours < 24 and 0 <= minutes < 60:
                return time(hours, minutes)
        except ValueError:
            pass
    
    # Try HMM format (3 digits)
    elif len(time_str) == 3 and time_str.isdigit():
        try:
            hours = int(time_str[0])
            minutes = int(time_str[1:])
            if 0 <= hours < 24 and 0 <= minutes < 60:
                return time(hours, minutes)
        except ValueError:
            pass
    
    return None
